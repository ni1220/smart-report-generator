"""
Native Excel chart generation engine using openpyxl.
Simplified version with robust error handling.
"""

import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.modules.ai_insight.models import PresentationPlan, SlideContent, ChartSpec

logger = logging.getLogger(__name__)


class ExcelGenerator:
    def __init__(self, plan: PresentationPlan):
        self._plan = plan
        self._wb = Workbook()
        self._wb.remove(self._wb.active)

    def generate(self) -> bytes:
        self._create_summary_sheet()
        self._create_ai_analysis_sheet()
        for slide in self._plan.slides:
            if slide.chart and slide.chart.categories and slide.chart.data_series:
                try:
                    self._create_chart_sheet(slide)
                except Exception as e:
                    logger.warning(f"Skip chart for page {slide.page_number}: {e}")
        buffer = BytesIO()
        self._wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _create_ai_analysis_sheet(self):
        """Create a sheet with AI analysis reasoning for each slide."""
        ws = self._wb.create_sheet("AI 分析思路")

        # Header
        headers = ["頁碼", "標題", "版面類型", "AI 驅動因素分析", "關鍵要點", "圖表說明"]
        header_font = Font(size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for i, slide in enumerate(self._plan.slides, 2):
            ws.cell(row=i, column=1, value=slide.page_number)
            ws.cell(row=i, column=2, value=slide.title)
            ws.cell(row=i, column=3, value=slide.layout_type)
            ws.cell(row=i, column=4, value=slide.insight_driver or "—")

            # Bullet points as combined text
            if slide.bullet_points:
                bullets = "\n".join(f"• {bp}" for bp in slide.bullet_points)
                ws.cell(row=i, column=5, value=bullets)
            else:
                ws.cell(row=i, column=5, value="—")

            # Chart info
            if slide.chart:
                chart_info = f"{slide.chart.chart_type}: {slide.chart.title}"
                if slide.chart.categories:
                    chart_info += f"\n維度: {', '.join(slide.chart.categories[:5])}"
                ws.cell(row=i, column=6, value=chart_info)
            else:
                ws.cell(row=i, column=6, value="—")

            # Wrap text for readability
            for col in range(1, 7):
                ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 30

        logger.info("Created AI analysis sheet")

    def _create_summary_sheet(self):
        ws = self._wb.create_sheet("執行摘要")
        ws["A1"] = "智匯數據簡報 — 數據摘要報告"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A3"] = self._plan.executive_summary
        ws["A5"] = "簡報頁面總覽"
        ws["A5"].font = Font(size=12, bold=True)
        for i, slide in enumerate(self._plan.slides, 6):
            ws.cell(row=i, column=1, value=slide.page_number)
            ws.cell(row=i, column=2, value=slide.title)
            ws.cell(row=i, column=3, value=slide.chart.chart_type if slide.chart else "—")

    def _create_chart_sheet(self, slide: SlideContent):
        chart_spec = slide.chart
        sheet_name = f"P{slide.page_number}_{chart_spec.title[:18]}"
        sheet_name = sheet_name.replace("/", "_").replace("\\", "_")[:31]
        ws = self._wb.create_sheet(sheet_name)

        # Write data
        ws.cell(row=1, column=1, value="項目")
        for col, series in enumerate(chart_spec.data_series, 2):
            ws.cell(row=1, column=col, value=series.name)

        for row_idx, category in enumerate(chart_spec.categories, 2):
            ws.cell(row=row_idx, column=1, value=category)
            for col_idx, series in enumerate(chart_spec.data_series, 2):
                if row_idx - 2 < len(series.values):
                    ws.cell(row=row_idx, column=col_idx, value=series.values[row_idx - 2])

        num_rows = len(chart_spec.categories) + 1
        num_cols = len(chart_spec.data_series) + 1

        # Create chart
        if chart_spec.chart_type == "pie":
            chart = PieChart()
        elif chart_spec.chart_type == "line":
            chart = LineChart()
        else:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "stacked" if chart_spec.chart_type == "stacked_bar" else "clustered"

        data_ref = Reference(ws, min_col=2, max_col=num_cols, min_row=1, max_row=num_rows)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.title = chart_spec.title
        chart.style = 10
        chart.width = 18
        chart.height = 12

        ws.add_chart(chart, f"A{num_rows + 2}")
        logger.info(f"Created chart: {chart_spec.title}")
